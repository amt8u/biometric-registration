"""Generate sample.xlsx matching the Blue Ridge Resident Registry schema.

Run:  python3 generate_sample.py
Requires: openpyxl  (pip install openpyxl)
"""
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT = "sample.xlsx"

REG_HEADERS = [
    "Timestamp", "BIN", "Type", "Type Label", "Tower", "Flat", "Residing",
    "Primary ID", "Full Name", "Mobile", "Email",
    "ID Document", "ID Number", "Agreement No", "Police Verif No",
    "Family Members", "Member IDs", "Member Names",
    "Drive Folder URL", "Registration Date", "Status",
]

MEM_HEADERS = [
    "Timestamp", "BIN", "Member ID", "Name", "ID Number", "Photo File", "ID Doc File",
]

DEL_HEADERS = [
    "Deleted Timestamp", "BIN", "Tower", "Flat", "Type", "Type Label",
    "Primary Name", "Primary Mobile", "Primary Email", "Member Count",
    "Registered On", "Reason", "Remarks", "Drive Folder URL", "Deleted By",
]

# ── Sample rows ──────────────────────────────────────────────────────────
REG_ROWS = [
    [
        datetime(2026, 4, 12, 10, 15), "TA-F0304-OWN", "OWN", "Owner (Residing)",
        "A", "304", "Yes",
        "TA-F0304-OWN-A", "Rajesh Kumar Sharma", "9820012345", "rajesh.sharma@example.com",
        "Aadhaar Card", "1234 5678 9012", "", "",
        2, "TA-F0304-OWN-B | TA-F0304-OWN-C", "Priya Sharma | Aarav Sharma",
        "https://drive.google.com/drive/folders/FAKE_OWN_RES_001",
        "12 Apr 2026", "ACTIVE",
    ],
    [
        datetime(2026, 4, 14, 16, 42), "TB-F1102-OWN-NR", "OWN", "Owner (Not Residing)",
        "B", "1102", "No",
        "TB-F1102-OWN-NR-A", "Sunita Mehta", "9900011223", "sunita.mehta@example.com",
        "PAN Card", "ABCDE1234F", "", "",
        0, "", "",
        "https://drive.google.com/drive/folders/FAKE_OWN_NR_002",
        "14 Apr 2026", "ACTIVE",
    ],
    [
        datetime(2026, 4, 18, 9, 5), "TC-F0501-TEN", "TEN", "Tenant",
        "C", "501", "N/A",
        "TC-F0501-TEN-A", "Mohammed Irfan Khan", "9012345678", "irfan.khan@example.com",
        "Driving Licence", "MH01 20210012345", "PUN-AGR-2026-4412", "PV-PUN-2026-8891",
        3,
        "TC-F0501-TEN-B | TC-F0501-TEN-C | TC-F0501-TEN-D",
        "Ayesha Khan | Zaid Khan | Sara Khan",
        "https://drive.google.com/drive/folders/FAKE_TEN_003",
        "18 Apr 2026", "ACTIVE",
    ],
    [
        datetime(2026, 4, 22, 19, 30), "TA-F0807-TEN", "TEN", "Tenant",
        "A", "807", "N/A",
        "TA-F0807-TEN-A", "Anjali Deshpande", "8877665544", "anjali.d@example.com",
        "Passport", "M1234567", "PUN-AGR-2026-4510", "PV-PUN-2026-8917",
        1, "TA-F0807-TEN-B", "Vikram Deshpande",
        "https://drive.google.com/drive/folders/FAKE_TEN_004",
        "22 Apr 2026", "DELETED",
    ],
]

MEM_ROWS = [
    [datetime(2026, 4, 12, 10, 15), "TA-F0304-OWN",    "TA-F0304-OWN-B",    "Priya Sharma",      "9876 5432 1098",   "Photo.jpg",     "ID_Document.pdf"],
    [datetime(2026, 4, 12, 10, 15), "TA-F0304-OWN",    "TA-F0304-OWN-C",    "Aarav Sharma",      "",                 "Photo.jpg",     ""],
    [datetime(2026, 4, 18, 9, 5),   "TC-F0501-TEN",    "TC-F0501-TEN-B",    "Ayesha Khan",       "FGHIJ5678K",       "Photo.jpg",     "ID_Document.pdf"],
    [datetime(2026, 4, 18, 9, 5),   "TC-F0501-TEN",    "TC-F0501-TEN-C",    "Zaid Khan",         "",                 "Photo.jpg",     ""],
    [datetime(2026, 4, 18, 9, 5),   "TC-F0501-TEN",    "TC-F0501-TEN-D",    "Sara Khan",         "",                 "Photo.jpg",     ""],
    [datetime(2026, 4, 22, 19, 30), "TA-F0807-TEN",    "TA-F0807-TEN-B",    "Vikram Deshpande",  "N7654321",         "Photo.jpg",     "ID_Document.pdf"],
]

DEL_ROWS = [
    [
        datetime(2026, 4, 25, 11, 0), "TA-F0807-TEN", "A", "807", "TEN", "Tenant",
        "Anjali Deshpande", "8877665544", "anjali.d@example.com", 1,
        "22 Apr 2026", "Tenant lease expired", "Vacated on 24 Apr 2026",
        "https://drive.google.com/drive/folders/FAKE_TEN_004",
        "manager@blueridge.co.in",
    ],
]


def write_sheet(ws, headers, rows):
    ws.append(headers)
    for r in rows:
        ws.append(r)
    # Header styling
    head_fill = PatternFill("solid", fgColor="1F3A5F")
    head_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"
    # Auto-ish column widths
    for col in range(1, len(headers) + 1):
        max_len = len(str(headers[col - 1]))
        for r in rows:
            v = r[col - 1]
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 45)


wb = Workbook()
ws1 = wb.active
ws1.title = "Registrations"
write_sheet(ws1, REG_HEADERS, REG_ROWS)

ws2 = wb.create_sheet("Family Members")
write_sheet(ws2, MEM_HEADERS, MEM_ROWS)

ws3 = wb.create_sheet("Deletion Log")
write_sheet(ws3, DEL_HEADERS, DEL_ROWS)

wb.save(OUT)
print(f"Wrote {OUT}")
